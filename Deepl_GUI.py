import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import io
import deepl
import os

# --- Page config ---
st.set_page_config(page_title="DeepL Excel Translator", layout="centered")
st.title("📄 DeepL Excel Translator")

# --- API key input ---
DEEPL_API_KEY = st.text_input("🔑 Enter your DeepL API Key", type="password")

# --- Upload file ---
uploaded_file = st.file_uploader("📁 Upload your Excel file (.xlsx or .xlsm)", type=["xlsx", "xlsm"])

if uploaded_file:
    st.markdown("### 🔧 Translation Settings")

    # Ask overwrite preference right after file is uploaded
    overwrite = st.radio("❓ Should we overwrite existing translations?", ["No", "Yes"], index=0)

    if not DEEPL_API_KEY:
        st.warning("Please enter your DeepL API key to continue.")
    else:
        # --- Read row 2 as header to detect columns ---
        df_headers = pd.read_excel(uploaded_file, header=1, nrows=0, engine="openpyxl")  # row 2 = header
        column_names = list(df_headers.columns)

        # Define all expected master content headers
        master_labels = [
            "Title - Master",
            "Product Description - Master",
            "Backend KW"
        ] + [f"Bullet Point {i} - Master" for i in range(1, 10)]

        # Filter only those present in the file
        valid_columns = [col for col in column_names if col in master_labels]

        if not valid_columns:
            st.warning("No Master Content columns found in row 2.")
        else:
            st.success("✅ Found the following Master Content columns:")
            st.write(valid_columns)

            # --- Language selection ---
            source_lang = st.text_input("🌍 Enter source language code (e.g., FR)", "")
            target_lang = st.text_input("🌍 Enter target language code (e.g., NL)", "")

            # --- Load workbook from memory ---
            in_memory_file = io.BytesIO(uploaded_file.getbuffer())
            wb = openpyxl.load_workbook(in_memory_file)
            ws = wb.active  # assumes first sheet

            wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            translation_done = False  # ✅ track if at least one translation was written

            # --- Helper to find column index by header (in row 2) ---
            def get_col_index(header_row, target_name):
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row, column=col).value == target_name:
                        return col
                return None

            # --- Translation function ---
            def translate_column(col_name):
                nonlocal translation_done
                st.write(f"🔁 Translating: **{col_name}**")
                col_idx = get_col_index(2, col_name)
                if not col_idx:
                    st.error(f"Column '{col_name}' not found.")
                    return

                translator = deepl.Translator(DEEPL_API_KEY)

                for row in range(3, ws.max_row + 1):  # content starts at row 3
                    source_cell = ws.cell(row=row, column=col_idx)
                    target_cell = ws.cell(row=row, column=col_idx + 1)

                    if source_cell.value:
                        if overwrite == "No" and target_cell.value not in (None, ""):
                            continue

                        try:
                            result = translator.translate_text(
                                str(source_cell.value),
                                source_lang=source_lang,
                                target_lang=target_lang
                            )
                            target_cell.value = result.text
                            target_cell.alignment = wrap_alignment
                            translation_done = True
                        except Exception as e:
                            target_cell.value = f"ERROR: {e}"
                            target_cell.alignment = wrap_alignment
                            translation_done = True

            # --- Column buttons ---
            st.markdown("### ✏️ Choose columns to translate:")
            col_buttons = {}
            for col in valid_columns:
                col_buttons[col] = st.button(f"Translate: {col}")

            translate_all = st.button("🌍 Translate All Content")

            # --- Execute translation ---
            for col in valid_columns:
                if col_buttons[col]:
                    translate_column(col)

            if translate_all:
                for col in valid_columns:
                    translate_column(col)

            # --- Download only if something was translated ---
            if translation_done:
                base_filename = os.path.splitext(uploaded_file.name)[0]
                output_filename = f"{base_filename}_translated.xlsx"

                output = io.BytesIO()
                wb.save(output)
                st.download_button(
                    label="📥 Download Translated File",
                    data=output.getvalue(),
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
